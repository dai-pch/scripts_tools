#!python3

import argparse
import re
import os
import os.path as path
import sys
from openpyxl import Workbook
from common_script import load_cfg, Executor, proc_meta
from tree import Tree


normal_indicator = "__normal$"

class PsudueData(object):
    def __str__(self):
        return 'not found'

class NormalData(PsudueData):
    def __init__(self, real_node, ref_path):
        self.real_node = real_node
        self.ref_path = ref_path

    @staticmethod
    def get_name(name):
        return name + '_normal'

def parse_contents(src, items_cfg, executor):
    res = Tree()
    for idx, item_cfg in enumerate(items_cfg):
        reg = item_cfg["regex"]
        search_iter = re.finditer(reg, src)
        for search_res in search_iter:
            item_name = item_cfg["name"]
            item_value = item_cfg["value"]
            executor.set_var("match", search_res.groups())
            item_name = executor.proc_embd_str(item_name)
            item_value = executor.proc_embd_str(item_value)
            value_type = "str"
            if "type" in item_cfg:
                value_type = item_cfg["type"]
            if value_type == 'int':
                item_value=int(item_value)
            elif value_type == 'float':
                item_value=float(item_value)
            # print(item_name)
            # print(item_value)
            node = res.root.add_child(item_value, item_name)
            if "normal-by" in item_cfg:
                assert(hasattr(item_value, '__truediv__'))
                normal_name = NormalData.get_name(item_name) 
                res.root.add_child(NormalData(node, item_cfg["normal-by"]), normal_name) 
    return res

def parse_name(regex, name):
    # regex = "cpu(\d+)-(LRU|T)_SWAP(-mram)?-(sp2k.+).txt$"
    sr = re.match(regex, name)
    if sr is None:
        raise ValueError('File name error.')
    return sr

def parse_file(file_path, cfg, executor, res=Tree()):
    f_dir, f_name = path.split(file_path)
    name_match = parse_name(cfg["file_name"], f_name)
    if not name_match:
        return res
    executor.set_var("file_name_match", name_match.groups())
    executor.set_var("item", name_match)
    f = open(file_path, 'r')
    c = parse_contents(f.read(), cfg["items"], executor)
    f.close()
    # save
    cur = res.root
    for hier in cfg["hierachy"]:
        hier_a = executor.proc_embd_str(hier)
        cur = cur.get_child_by_name(hier_a, set_default=True)
    cur.absorb_tree(c)
    # print(dic)
    # res.display()
    return res

def parse_folder(folder_path, cfg, executor, res=Tree()):
    for root, dirs, files in os.walk(folder_path):
        for f in files:
            f_path = path.join(root, f)
            try:
                res = parse_file(f_path, cfg, executor, res)
            except ValueError:
                continue
    # res = Tree.from_dict(res)
    return res

def post_proc(data):
    proc_normal(data)
    return data

def proc_normal(data):
    # find normal items
    normal_items = []
    def get_normal_node(node):
        nonlocal normal_items
        if isinstance(node.data, NormalData):
            normal_items.append(node)
    data.root.depth_first_traverse(func_leaf=get_normal_node)
    # proc
    for node in normal_items:
        real_node, ref_path = node.data.real_node, node.data.ref_path
        cur = node
        ref = None
        while cur:
            ref = cur.get_child_by_path(ref_path)
            if ref:
                break
            cur = cur.get_parent()
        rem_path = Tree.extract_path(cur, real_node)
        if len(rem_path) < len(ref_path):
            print("Warning: Can not find normal ref ", str(ref_path), " for ", node.name)
            continue
        rem_path = rem_path[len(ref_path):]
        if not ref:
            print("Warning: Can not find normal ref ", str(ref_path), " for ", node.name)
            continue
        ref = ref.get_child_by_path(rem_path)
        if not ref:
            print("Warning: Can not find normal ref ", str(ref_path), " for ", node.name)
            continue
        node.data = float(real_node.data) / float(ref.data) if 0 !=ref.data else (0 if float(real_node.data) == 0 else float(Nan))

def gen_blk(v, row_n, ws):
    # scan col item
    col_items = v.get_next_n_nodes(v.get_level()-1)
    col_items = list(set([n.name for _, n in col_items]))
    # add col name
    col_items.sort()
    for idy, col_item in enumerate(col_items):
        ws.cell(row=idy+row_n+1, column=1).value = col_item
        
    # add blk
    gen_row_blk(v, row_n, 1, 2, col_items, ws)

def gen_row_blk(root, row_n, r_s, c_s, item_list, ws):
    # row header
    if row_n == 0: # leaf
        for node in root.get_children():
            idy = item_list.index(node.name)
            ws.cell(row=idy+r_s, column=c_s).value = node.data
        return 1
    # not leaf
    c_n = c_s
    items = list(root.get_children())
    items.sort(key=lambda x:x.name)
    for n in items:
        w = gen_row_blk(n, row_n-1, r_s+1, c_n, item_list, ws)
        ws.merge_cells(start_row=r_s, end_row=r_s, start_column=c_n, end_column=c_n+w-1)
        ws.cell(row=r_s, column=c_n).value = n.name
        c_n += w
    return c_n - c_s

def gen_ws(k, v, row_n, wb):
    ws = wb.create_sheet(title=k)
    gen_blk(v, row_n, ws)

def gen_wb(name, data, ws_n, row_n):
    wb = Workbook()
    ws_data = data.get_next_n_nodes(ws_n)
    _ = list(map(lambda tp: gen_ws("-".join(tp[0][1:]), tp[1], row_n, wb), ws_data))
    return (name, wb)

def gen_tb(data, d_name, wb_n=0, ws_n=1):
    data.root.name = d_name
    col_n = 1
    assert(col_n + wb_n + ws_n <= data.tree_depth())
    row_n = data.tree_depth() - col_n - wb_n - ws_n - 1
    wb_data = data.root.get_next_n_nodes(wb_n)
    wbs = list(map(lambda tp: gen_wb("-".join(tp[0]), tp[1], ws_n, row_n), wb_data))
    return wbs

def write_tbs(tbs, folder):
    for name, tb in tbs:
        tb.save(path.join(folder, name + ".xlsx"))

def get_argparser():
    parser = argparse.ArgumentParser()
    parser.add_argument('dir', type=str)
    parser.add_argument('-L', '--config-path', type=list)
    parser.add_argument('-c', '--config', type=str, default="parser.cfg")
    parser.add_argument('-o', '--default-name', type=str, default="result")
    parser.add_argument('--run-begin', default="<%")
    parser.add_argument('--run-end', default="%>")
    return parser

def run():
    parser = get_argparser()
    args = parser.parse_args()
    cfg = load_cfg(args.config, args.config_path)
    executor = Executor()
    proc_meta(args, cfg["meta"], executor)
    cfg = cfg["config"]
    data = parse_folder(args.dir, cfg, executor)
    data = post_proc(data)
    xslx_cfg = cfg["xslx"]
    tbs = gen_tb(data, args.default_name, xslx_cfg["book"], xslx_cfg["sheet"])
    write_tbs(tbs, sys.argv[1])

if __name__ == "__main__":
    run()
    # print(res)

